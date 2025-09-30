import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def validate_hours(row):
    les_week_th_col = 'Εβδ. Ωρες Θεωρίας'
    les_tot_th_col = 'Συνολ. Ώρες Θεωρίας'
    les_week_lab_col = 'Εβδ. Ωρες Εργαστ.'
    les_tot_lab_col = 'Συνολ. Ώρες Εργαστ.'

    week_th = row[les_week_th_col]
    week_lab = row[les_week_lab_col]
    tot_th = row[les_tot_th_col]
    tot_lab = row[les_tot_lab_col]
    err =""
    if week_th == 0 and week_lab == 0 and tot_th == 0 and tot_lab == 0:
        return "Όλες οι ώρες μαθήματος είναι μηδέν"
    if week_th > tot_th:
        err = "Ο αριθ εβδ. ωρ. θεωρίας είναι μεγαλύτερος από τον συνολικό,"

    if week_lab > tot_lab:
        err = "Ο αριθ εβδ. ωρ. εργ. είναι μεγαλύτερος από τον συνολικό,"

    err = err.strip()
    return err

def validate_field(field, value):
    yesno = ["ΝΑΙ", "ΟΧΙ"]
    les_years = [1,2]
    assignments = ['Α','Β','Γ','Δ']
    err =""
    
    if field == "Μόνο με δευτεροβάθμια":
        if value not in yesno: err += f"{field}: Μη έγκυρη τιμή,"
    if field == "Έτος Μαθήματος": 
        if value not in les_years: err += f"{field}: Μη έγκυρη τιμή,"
    if field == "Ανάθεση": 
        if value not in assignments: err += f"{field}: Μη έγκυρη τιμή,"
    if field == "Κωδ. Ειδικ. Εκπαιδευτικού": 
        if value not in range(1,209): err += f"{field}: Μη έγκυρη τιμή,"

    err = err.strip()
    return err

def validate_df():
    global spc,lc, total_errors

    sp_col = 'Όνομα Ειδικότητας'
    deyt_col = 'Μόνο με δευτεροβάθμια'
    les_col = 'Τίτλος Μαθήματος'
    les_year_col = 'Έτος Μαθήματος'
    les_week_th_col = 'Εβδ. Ωρες Θεωρίας'
    les_tot_th_col = 'Συνολ. Ώρες Θεωρίας'
    les_week_lab_col = 'Εβδ. Ωρες Εργαστ.'
    les_tot_lab_col = 'Συνολ. Ώρες Εργαστ.'
    t_sp_id_col = 'Κωδ. Ειδικ. Εκπαιδευτικού'
    t_sp_assign_col = 'Ανάθεση'
    t_sp_comm_col = 'Προαπαιτούμενα'

    
    req_cols = [sp_col,deyt_col,les_col,les_year_col,les_week_th_col,les_week_th_col,les_tot_th_col,les_week_lab_col,les_tot_lab_col,t_sp_id_col,t_sp_assign_col]
    
    specs = {}
    lessons = {}
    #les_specs = {}
    les_specs = []
    for i,row in df.iterrows():
        has_err = False
        sp = row[sp_col]
        deyt = row[deyt_col]
        les = row[les_col].strip()
        year = row[les_year_col]
        week_th = row[les_week_th_col]
        week_lab = row[les_week_lab_col]
        tot_th = row[les_tot_th_col]
        tot_lab = row[les_tot_lab_col]
        #les_type = row[les_type_col]
        t_sp = row[t_sp_id_col]
        #exam = row[les_exam_col]
        assign = row[t_sp_assign_col]
        comments = row[t_sp_comm_col]
        err = ""
        
        for col in req_cols:
           if pd.isna(row[col]):
               err += f"{col}: Υποχρεωτικό πεδίο,"
           else : err += validate_field(col, row[col])
           if err != "" and not has_err:
               has_err = True
               total_errors += 1 
           

        hour_err = validate_hours(row)
        if hour_err != "":
            err += hour_err
            if not has_err:
                has_err = True
                total_errors += 1 

        if pd.isna(sp) or pd.isna(les) or pd.isna(assign):
            if err != None:
                if err[-1] == ",": err = err[:-1]  
                errors[i] = err
                print(f"{i}: errors: {err}")
            continue

        les_key = sp + "_" + les + "_" + str(year)
        les_spec_key = les_key + "_" + str(t_sp) + "_" + assign

        if not sp in specs.keys():
            specs[sp] = spc
            spc+=1
        if not les_key in lessons.keys():
            lessons[les_key] = lc
            lc+=1
        if not les_spec_key in les_specs:
            les_specs.append(les_spec_key)
        else:
            #print(f"row: {i}: Η ειδικότητα υπάρχει ήδη για το μάθημα. {les_spec_key}")
            if err == None: err = ""
            err += f"row: {i}: Η ανάθεση υπάρχει ήδη για αυτήν την ειδικότητα. {les_spec_key},"
            if not has_err:
                has_err = True
                total_errors += 1 
            #exit()
        err = err.strip()
        if err != "":
            errors[i] = err
            if err[-1] == ",": err = err[:-1]  
            #df.loc[i, 'ΣΦΑΛΜΑΤΑ'] = errors[i]
            #print(f"{i}: errors: {err}")
        
    print("total rows with errors: ", total_errors)

def highlight_errors_and_save():
    #fa0a22
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow highlight
    #highlight_fill = PatternFill(start_color="fa0a22", end_color="fa0a22", fill_type="solid")  # Yellow highlight
    #df['ΣΦΑΛΜΑΤΑ'] = ""
    wb = load_workbook(source)
    ws = wb.active
    for i, row in df.iterrows():
        if i not in errors.keys(): continue
        for cell in ws[i + 2]:
            cell.fill = highlight_fill
        ws.cell(row=i+2, column=ws.max_column).value = errors[i]    
        #df.loc[i, 'ΣΦΑΛΜΑΤΑ'] = errors[i]
    wb.save("errors/amea_ath_lessons.xlsx")
  
def fill_lists():
    global spc,lc
    sp_col = 'Όνομα Ειδικότητας'
    deyt_col = 'Μόνο με δευτεροβάθμια'
    les_col = 'Τίτλος Μαθήματος'
    les_year_col = 'Έτος Μαθήματος'
    les_week_th_col = 'Εβδ. Ωρες Θεωρίας'
    les_tot_th_col = 'Συνολ. Ώρες Θεωρίας'
    les_week_lab_col = 'Εβδ. Ωρες Εργαστ.'
    les_tot_lab_col = 'Συνολ. Ώρες Εργαστ.'
    #les_type_col = 'Τύπος Μαθήματος'
    #les_exam_col = 'Εξεταστέο'
    t_sp_id_col = 'Κωδ. Ειδικ. Εκπαιδευτικού'
    t_sp_assign_col = 'Ανάθεση'
    t_sp_comm_col = 'Προαπαιτούμενα'

    
    specs = {}
    lessons = {}
    #les_specs = {}
    les_specs = []
    for i,row in df.iterrows():
        sp = row[sp_col]
        deyt = row[deyt_col]
        les = row[les_col].strip()
        year = row[les_year_col]
        week_th = row[les_week_th_col]
        week_lab = row[les_week_lab_col]
        tot_th = row[les_tot_th_col]
        tot_lab = row[les_tot_lab_col]
        #les_type = row[les_type_col]
        t_sp = row[t_sp_id_col]
        #exam = row[les_exam_col]
        assign = row[t_sp_assign_col]
        comments = row[t_sp_comm_col]

        if pd.isna(comments): comments=None
        else: comments = comments.strip()

        if pd.isna(sp):
            print(f"row: {i}: Μη έγκυρο όνομα ειδικότητας: {i} - {sp} - {les}")
            exit()
        else: sp = sp.strip()

        if assign not in ['Α','Β','Γ','Δ']:
            print(f"row: {i}: Μη έγκυρος τύπος ανάθεσης: {i} - {sp} - {les}")
            exit()

        les_key = sp + "_" + les + "_" + str(year)
        les_spec_key = les_key + "_" + str(t_sp) + "_" + assign

        if not sp in specs.keys():
            specs[sp] = spc
            if deyt  not in ['ΝΑΙ', 'ΟΧΙ']:
                print(f"row: {i}: Μη έγκυρος τύπος δευτεροβάθμιας εκπ. {les_key}")
                exit()
            else: deyt = deyt.strip()
            # id, name,dypa_institution_type_id, edu_id, higher_education_only
            higher_ed = 1 if deyt == 'ΝΑΙ' else 0
            specs_val.append(f"\n({spc},'{sp}',5,95,{higher_ed},now())")
            spc+=1
        if not les_key in lessons.keys():
            lessons[les_key] = lc
            
            if (week_th ==0 and tot_th !=0) or (week_th !=0 and tot_th ==0):
                print(f"row: {i}: Οι ώρες θεωρίας είναι λάθος. {les_spec_key}")
                exit()
            if (week_lab ==0 and tot_lab !=0) or (week_lab !=0 and tot_lab ==0):
                print(f"row: {i}: Οι ώρες εργαστηρίου είναι λάθος. {les_spec_key}")
                exit()
            has_th = 1 if week_th > 0 and tot_th > 0 else 0
            has_lab = 1 if week_lab > 0 and tot_lab > 0 else 0 
            lesson_type = 3 if has_th and has_lab else 1 if has_th else 2
            exam_val = 1
            # id,descr,student_specialty_id,hours_th,hours_erg,has_erg,has_theory,lesson_type,year,,examinable10,is_active,total_hours_erg,total_hours_th,period_num,has_special_deg,created_at16
            les_val.append(f"\n({lc},'{les}',{specs[sp]},{week_th},{week_lab},{has_lab},{has_th},{lesson_type},{year},{exam_val},1,{tot_lab},{tot_th},{year},0,now())")
            lc+=1
        if not les_spec_key in les_specs:
            les_specs.append(les_spec_key)
            com = "NULL" if comments is None else comments.strip()
            #anathesi,lesson_id,student_specialty_id,comments_docs
            les_specs_val.append(f"\n('{assign}',{lessons[les_key]},{t_sp},{com})")
        else:
            print(f"row: {i}: Η ειδικότητα υπάρχει ήδη για το μάθημα. {les_spec_key}")
            exit()

def make_queries():
    insert_specs = f"INSERT INTO student_specialties (id, name, dypa_institution_type_id,edu_id, higher_education_only, created_at) VALUES {','.join(specs_val)};"
    alter_specs = f"ALTER TABLE student_specialties AUTO_INCREMENT = {spc};"

    insert_lessons = f"INSERT INTO aa_lessons (id,descr,student_specialty_id,hours_th,hours_erg,has_erg,has_theory,lesson_type,year,examinable,is_active,total_hours_erg,total_hours_th,period_num,has_special_deg,created_at) VALUES {','.join(les_val)};"
    alter_lessons = f"ALTER TABLE aa_lessons AUTO_INCREMENT = {lc};"

    insert_lesson_specs = f"INSERT INTO aa_lesson_specialties (anathesi, lesson_id, specialty_id,comments_docs) VALUES {','.join(les_specs_val)};"
    #alter_lesson_specs = f"ALTER TABLE aa_lesson_specialties AUTO_INCREMENT = {len(les_specs_val) +1 };"


    return f"{insert_specs}\n{alter_specs}\n\n{insert_lessons}\n{alter_lessons}\n\n{insert_lesson_specs}\n\n"

# db current last student specialty id
spc=146
# db current last lesson id
lc =40
error_source = "errors/amea_ath_lessons.xlsx"
source = "data/amea_ath_lessons.xlsx"
specs_val,les_val,les_specs_val = [],[],[]
df = pd.read_excel(source)


total_errors = 0
errors = {}
validate_df()
if total_errors > 0:
    #df.to_excel(error_source, index=False)
    print("Errors -----------")
    for i in errors.keys():
        print(f"{i}: {errors[i]}")
    highlight_errors_and_save()
    #print(df.head())

# fill_lists()
# res = make_queries()
# with open("out/amea_ath_spec_lessons.sql", "w", encoding="utf8") as f: f.write(res)